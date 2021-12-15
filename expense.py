#!/usr/bin/python3

# NO-FRILS text only expense tracker for sole proprietorship
# 
# See README.md for documentation and usage instruction
# at github: https://github.com/mbe2014/Expense
#       
# Copyright (c) 2021, Moshe Ben-Ezra
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# 1. Redistributions of source code must retain the above copyright notice, this
#    list of conditions and the following disclaimer.
#
# 2. Redistributions in binary form must reproduce the above copyright notice,
#    this list of conditions and the following disclaimer in the documentation
#    and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
# SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
# OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
# OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.

import os
import sys
import openpyxl as xl
import json
import shutil
import hashlib

# Expense categories from Tax Form 1040 Schedule C 
# Profit and Loss from business (sole proprietorship)
# each category has a list that contages signatures (payee:description) of
# transactions that fit this category. The list is updated dynamically and
# saved in "categories.json" when changes are saved
categories = {"Advertising":[], 
              "Car and truck expenses":[],
              "Commissions and fees":[],
              "Contract labor":[],
              "Depletion":[],
              "Depreciation and section 179 expense deduction":[],
              "Employee benefit programs":[],
              "Insurance":[], "Interest (Mortgage)":[],
              "Interest (Other)":[],
              "Legal and professional services":[],
              "Office expense and software":[], 
              "Pension and profit-sharing plans":[],
              "Rent or lease (Vehicle and machinery)":[],
              "Rent or lease (other business property)":[],
              "Repairs and maintenance":[],
              "Supplies":[], 
              "Taxes and licenses":[],
              "Travel and meals (travel)":[],
              "Travel and meals (meals)":[],
              "Utilities":[],
              "Wages":[],
              "Other expenses":[],
              "Not Applicable":[]}

# Field titles orgenized by priority. category tite (catg_title) is not
# currently used is is present for future use
date_title   = ["Date", "Transaction Date", "Trans. Date", "Posted Date"]
ref_title    = ["Reference", "Reference ID", "Reference Number", "Transaction ID"]
payee_title  = ["Payee","Name"]
desc_title   = ["Description", "Memo", "Subject"]
catg_title   = []
sum_title    = ["Amount","Net", "Sum"]

# List to save transaction that are accepted, skipped, or duplicate to their log file
accept_log = []
skip_log   = []
dup_log    = []

# List of references in existig and added transactions
# to prevent enteruing duplicate transactions
reference_list = []

# Initiaze the program. When reset is True init() will
# initialize the repository as well REMOVING ALL FILES
# and initialize the repo and categories.
# othewise init() load the repo and categories form their files

def init(repo, fname, reset):
    global wb
    global ws
    global categories
    global key_list
    global sdir

    wb_name   = repo + "/" + fname
    cat_name  = repo + "/" + "categories.json"
    sdir      = repo + "/" + "STATEMENTS"

    if reset == True:
        if os.path.isdir(repo):
            ans = input(f"\nAre you sure you want to remove all files in {repo}? [y/N]:")
            if ans != "y":
                print("Abort..\n")
                sys.exit(0)

        shutil.rmtree(repo, ignore_errors=True)
        os.mkdir(repo)
        os.mkdir(sdir)       
        os.system(f"git init {repo}")

    key_list = []
    for x in categories.keys():
        key_list.append(x)

    if not os.path.isfile(wb_name):
        wb = xl.Workbook() 
        ws = wb.active
        ws['A1'] = "Date"
        ws['B1'] = "Reference"
        ws['C1'] = "Payee"
        ws['D1'] = "Category"
        ws['E1'] = "Description"
        ws['F1'] = "Amount"
    else:
        wb = xl.load_workbook(wb_name)
        ws = wb.active
        # init reference list to detect duplicates
        for x in range (2, ws.max_row+1):
            ref = ws.cell(row=x,column=2).value
            val = ws.cell(row=x,column=6).value
            if ref != "na":
                reference_list.append(ref + ":" + str(val))

    if os.path.isfile(cat_name):
        with open(cat_name, 'r') as f:
            categories = json.load(f)

# save the updated repo, categories and log files
# Save is done at the end to allow the user to abort
# and not save changes in case something went wrong

def save_all(repo,fname, inp_name):
    wb_name =  repo + "/" + fname
    cat_name = repo + "/" + "categories.json"
    acc_log_name = repo + "/" + "accept_entries.csv"
    skp_log_name = repo + "/" + "skipped_entries.csv"
    dup_log_name = repo + "/" + "duplicate_entries.csv"

    wb.save(wb_name)

    with open(cat_name, 'w') as f:
        json.dump(categories, f, indent=2) 

    with open(acc_log_name, 'a') as f:
        f.write("\n"+inp_name+"\n\n")
        for x in accept_log:
            f.write(f"{x[0]},{x[1]},{x[2]},{x[3]},{x[4]},{x[5]}\n")

    with open(skp_log_name, 'a') as f:
        f.write("\n"+inp_name+"\n\n")
        for x in skip_log:
            f.write(f"{x[0]},{x[1]},{x[2]},{x[3]},{x[4]},{x[5]}\n")

    with open(dup_log_name, 'a') as f:
        f.write("\n"+inp_name+"\n\n")
        for x in dup_log:
            f.write(f"{x[0]},{x[1]},{x[2]},{x[3]},{x[4]},{x[5]}\n")

# process a transaction file
# this is the main function that does all the work
# changes are kept in memory until saved

def process(fname):
    print ("\033[H\033[J")
    print ("Processing:", fname)

    # step 1. load input transactions and identify fields
    # ---------------------------------------------------
    inb  = xl.load_workbook(fname, read_only=True)
    ins = inb.active

    rows = ins.max_row
    cols = ins.max_column
    print ("Input sheet size:",rows, cols,"\n")

    date_col  = None
    ref_col   = None
    payee_col = None
    desc_col  = None
    sum_col   = None

    print ("Data fields")
    print ("-----------")

    for x in range (1,cols+1):
        val = ins.cell(row=1,column=x).value
        if date_col == None and val in date_title:
            date_col = x
            print (f'Date in column:        {x:2} Title: "{val}"')

        if ref_col == None and val in ref_title:
            ref_col = x
            print (f'Reference in column:   {x:2} Title: "{val}"')

        if payee_col == None and val in payee_title:
            payee_col = x
            print (f'Payee in column:       {x:2} Title: "{val}"')
       
        if  desc_col == None and val in desc_title:
            desc_col = x
            print (f'Description in column: {x:2} Title: "{val}"')

        if sum_col == None and val in sum_title:
            sum_col = x
            print (f'Sum in column:         {x:2} Title: "{val}"')


    ans = input("\nPlease confirm [Y/n]:")
    if ans == "n":
        print("Abort..\n")
        sys.exit(0)

    if date_col == None or sum_col == None:
        print ("Date and amount fields are mandatory")
        print("Abort..\n")
        sys.exit(0)

    if desc_col == None  and payee_col == None: 
        print ("Description or Payee fields are mandatory")
        print("Abort..\n")
        sys.exit(0)

    # Step 2. process transactions
    # ----------------------------
    for x in range (2,rows+1):
        date  = ins.cell(row=x,column=date_col).value
        amnt  = ins.cell(row=x,column=sum_col).value

        if date == None or amnt == None:
            print ("Date and amount entries are mandatory")
            print("Abort..\n")
            sys.exit(0)

        ref   = None
        payee = None
        desc  = None
        cat   = None

        if ref_col != None: 
            ref   = ins.cell(row=x,column=ref_col).value        
        if payee_col != None:
            payee = ins.cell(row=x,column=payee_col).value
        if desc_col != None:
            desc  = ins.cell(row=x,column=desc_col).value

        if ref == None:
            ref = "na"
        if payee == None:
            payee = "na"
        if desc == None:
            desc = "na"

        # Step 2.1 look for existing catgory
        # -------------------------
        print("payee:",payee)
        print("desc:",desc)
        pattern = payee + ":" + desc
        for k in key_list:
            pat_list = categories[k]
            if pattern in pat_list:
                cat = k
                print (f"Pattern {pattern} is category {cat}")
                break
        
        # Step 2.2 if not found ask user to categorize transaction and remember
        # ---------------------------------------------------------------------
        if cat == None:
            print("\033[H\033[J")
            for x in range (1, len(key_list)):
                print (x, key_list[x])

            ans = 0
            while ans < 1 or ans > len(key_list):
                ans = input ("\nSelect category number for: " + pattern + ": ")
                ans = int(ans)

            cat = key_list[ans]
            categories[cat].append(pattern)


        # Step 2.3 prepare new entry and update repo and logs as needed
        # -------------------------------------------------------------    
        entry = (date, ref, payee, cat, desc, abs(amnt))

        print("\n",entry)

        if (cat == "Not Applicable"):
            print("\nNot Applicable entry detected..\n")
            ans = input("\nSkip NA entry? [Y/n]:")
            if ans != "n":
                skip_log.append(entry)
                continue

        ref_sig = ref + ":" + str(amnt)
        if ref != "na" and ref_sig in reference_list:
            print("\nDuplicate detected..\n")
            ans = input("\nSkip duplicate entry? [Y/n]:")
            if ans != "n":
                dup_log.append(entry)
                continue  

        ans = input("\nPlease confirm [Y/n]:")
        if ans == "n":
            print("skipped..\n")
            skip_log.append(entry)
            continue

        ws.append(entry)
        reference_list.append(ref)
        accept_log.append(entry)


def usage():
        print(f"Usage: {sys.argv[0]} init folder-name")
        print(f"       {sys.argv[0]} import folder-name file.xlsx")


def sha256(fname):
    with open(fname,"rb") as f:
        bytes = f.read() # read entire file as bytes
        hash_str = hashlib.sha256(bytes).hexdigest();
        return hash_str


if __name__ == "__main__":

    if len(sys.argv) < 3:
        usage()
        sys.exit(0)

    cmd  = sys.argv[1]
    repo = sys.argv[2]

    if len(sys.argv) == 4:
        infile = sys.argv[3]

    if cmd == "init" and len(sys.argv) !=3:
        usage()
        sys.exit(0)

    if cmd == "import" and len(sys.argv) != 4:
        usage()
        sys.exit(0)
  
    fname = "workbook.xlsx"
    path = repo + "/" + fname
    
    git_comment = ""
    if cmd == "init":
        git_comment = "Initial Commit"
        init(repo,fname,True);

    elif cmd == "import":
        if not os.path.isdir(repo) or not  os.path.isfile(path):
            print(f"Folder {repo} not initialized")
            print(f'run {sys.argv[0]} init {repo}" to initialize')
            sys.exit(0)

        git_comment = "Import: " + infile
        init(repo,fname,False);
        process(infile)
        
    else:
        usage()
        sys.exit(0)

    ans = input("\nSave changes? [Y/n]:")
    if ans == "n":
        print("Abort..\n")
        sys.exit(0)

    save_all(repo,fname,git_comment)
    
    if cmd == "import":
        sha_name = repo + "/STATEMENTS/" +  sha256(infile) + "_" + os.path.basename(infile)
        shutil.copyfile(infile, sha_name)
    
    os.chdir(repo)
    os.system("git add workbook.xlsx categories.json accept_entries.csv skipped_entries.csv duplicate_entries.csv STATEMENTS")
    os.system(f'git commit -m "{git_comment}"')
