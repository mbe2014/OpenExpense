#!/usr/bin/python3

# Tax repot for tyhe NO-FRILS text only expense tracker for sole proprietorship
# 
# See README.md for documentation and usage instruction
# at github: https://github.com/mbe2014/Expense
#       
# Copyright (c) 2021, Moshe Ben-Ezra
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# 1. Redistributions of source code must retain the above copyright notice, this
#    list of conditions and the following disclaimer.
#
# 2. Redistributions in binary form must reproduce the above copyright notice,
#    this list of conditions and the following disclaimer in the documentation
#    and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
# SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
# OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
# OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.

import os
import sys
import openpyxl as xl
import json
import shutil
import hashlib

# Expense categories from Tax Form 1040 Schedule C 
# Profit and Loss from business (sole proprietorship)

sum_cat = { "Advertising":0, 
            "Car and truck expenses":0,
            "Commissions and fees":0,
            "Contract labor":0,
            "Depletion":0,
            "Depreciation and section 179 expense deduction":0,
            "Employee benefit programs":0,
            "Insurance":0, 
            "Interest (Mortgage)":0,
            "Interest (Other)":0,
            "Legal and professional services":0,
            "Office expense":0, 
            "Pension and profit-sharing plans":0,
            "Rent or lease (Vehicle and machinery)":0,
            "Rent or lease (other business property)":0,
            "Repairs and maintenance":0,
            "Supplies":0, 
            "Taxes and licenses":0,
            "Travel and meals (travel)":0,
            "Travel and meals (meals)":0,
            "Utilities":0,
            "Wages":0,
            "Other expenses":0,
            "Not Applicable":0 }


def process(fname):

    # step 1. load input transactions and identify fields
    # ---------------------------------------------------
    inb  = xl.load_workbook(fname, read_only=True)
    ins = inb.active

    rows = ins.max_row
    cols = ins.max_column
    print ("Input sheet size:",rows, cols,"\n", file=sys.stderr)

    # Step 2. process table
    # ----------------------------
    sum_all = 0
    for x in range (2,rows+1):
        cat  = ins.cell(row=x,column=4).value
        amnt = ins.cell(row=x,column=6).value
        # print (cat," : ",amnt)
        if cat in sum_cat:
            sum_cat[cat] = sum_cat[cat] + amnt
            sum_all = sum_all + amnt
        else:
            print(f"Error: {cat} not in IRS caterories", file=sys.stderr)

    # Step 3. print report
    # ----------------------------
    for x in sum_cat:
        if sum_cat[x] != 0:
            print("{:46s}{:10.2f}".format(x,sum_cat[x]))
    print("")
    print("{:46s}{:10.2f}".format("Sum All",sum_all))

def usage():
        print(f"Usage: {sys.argv[0]} folder-name")

if __name__ == "__main__":

    if len(sys.argv) != 2:
        usage()
        sys.exit(0)

    repo = sys.argv[1]
  
    fname = "workbook.xlsx"
    path = repo + "/" + fname
    process(path)    
